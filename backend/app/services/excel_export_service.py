import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill("solid", fgColor="0F172A")
SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11, bold=False, color="111827")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="475569")
CARD_TITLE_FONT = Font(name="Calibri", size=10, bold=True, color="475569")
CARD_VALUE_FONT = Font(name="Calibri", size=18, bold=True, color="0F172A")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
NO_BORDER = Border(
    left=Side(style=None),
    right=Side(style=None),
    top=Side(style=None),
    bottom=Side(style=None),
)
CURRENCY_FORMAT = '#,##0.00 [$EUR]'


def adjust_column_width(ws, min_width: int = 12, max_width: int = 32):
    for column in ws.columns:
        cells = list(column)
        if not cells:
            continue
        max_length = 0
        for cell in cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(max(max_length + 2, min_width), max_width)


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=11, bold=False)
        cell.border = NO_BORDER


def generate_excel_from_df(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        style_header(worksheet)
        adjust_column_width(worksheet)
        worksheet.sheet_view.showGridLines = False

        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Calibri", size=11, bold=False)
                cell.border = NO_BORDER

    buffer.seek(0)
    return buffer.read()


def _set_cell(cell, value, *, font=BODY_FONT, fill=None, alignment=None, border=THIN_BORDER, number_format=None):
    cell.value = value
    cell.font = font
    cell.border = border
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def _merge_title(ws, title: str, subtitle: str):
    ws.merge_cells("B2:H3")
    title_cell = ws["B2"]
    title_cell.value = title
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B4:H4")
    subtitle_cell = ws["B4"]
    subtitle_cell.value = subtitle
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")


def _metric_card(ws, start_col: int, start_row: int, title: str, value, accent: str):
    fill = PatternFill("solid", fgColor=accent)
    for row in range(start_row, start_row + 3):
        for col in range(start_col, start_col + 2):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = NO_BORDER

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=start_col + 1)
    _set_cell(ws.cell(start_row, start_col), title, font=CARD_TITLE_FONT, fill=fill, border=NO_BORDER)
    _set_cell(ws.cell(start_row + 1, start_col), value, font=CARD_VALUE_FONT, fill=fill, border=NO_BORDER)


def _write_table(ws, start_row: int, start_col: int, headers: list[str], rows: list[list], title: str | None = None):
    current_row = start_row
    if title:
        ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + len(headers) - 1)
        _set_cell(
            ws.cell(current_row, start_col),
            title,
            font=Font(name="Calibri", size=12, bold=True, color="0F172A"),
            fill=SECTION_FILL,
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        current_row += 1

    for index, header in enumerate(headers, start=start_col):
        _set_cell(
            ws.cell(current_row, index),
            header,
            font=HEADER_FONT,
            fill=HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center"),
        )
    current_row += 1

    if not rows:
        rows = [["Sin datos disponibles"] + [None] * (len(headers) - 1)]

    for row in rows:
        for offset, value in enumerate(row, start=start_col):
            cell = ws.cell(current_row, offset)
            _set_cell(cell, value)
            if isinstance(value, (int, float)) and offset not in (start_col,):
                cell.number_format = CURRENCY_FORMAT if "importe" in str(headers[offset - start_col]).lower() or "saldo" in str(headers[offset - start_col]).lower() else "0"
        current_row += 1

    return current_row - 1


def _create_bar_chart(ws, min_col: int, min_row: int, max_row: int, title: str, position: str, category_col: int = 1):
    if max_row <= min_row:
        return
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.height = 8
    chart.width = 16
    chart.y_axis.title = "EUR"
    chart.x_axis.title = "Periodo"
    data = Reference(ws, min_col=min_col, min_row=min_row, max_col=min_col, max_row=max_row)
    categories = Reference(ws, min_col=category_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    ws.add_chart(chart, position)


def _create_line_chart(ws, data_columns: list[int], min_row: int, max_row: int, title: str, position: str, category_col: int = 1):
    if max_row <= min_row:
        return
    chart = LineChart()
    chart.title = title
    chart.style = 13
    chart.height = 8
    chart.width = 18
    chart.y_axis.title = "EUR"
    chart.x_axis.title = "Semana"
    categories = Reference(ws, min_col=category_col, min_row=min_row + 1, max_row=max_row)
    for column in data_columns:
        data = Reference(ws, min_col=column, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, position)


def _create_pie_chart(ws, label_col: int, value_col: int, min_row: int, max_row: int, title: str, position: str):
    if max_row <= min_row:
        return
    chart = PieChart()
    chart.title = title
    chart.height = 7
    chart.width = 10
    labels = Reference(ws, min_col=label_col, min_row=min_row + 1, max_row=max_row)
    data = Reference(ws, min_col=value_col, min_row=min_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    ws.add_chart(chart, position)


def _format_currency_columns(ws, columns: list[int], start_row: int, end_row: int):
    for row in range(start_row, end_row + 1):
        for col in columns:
            ws.cell(row=row, column=col).number_format = CURRENCY_FORMAT


def generate_dashboard_excel(payload: dict) -> bytes:
    stats = payload.get("stats", {})
    treasury = payload.get("treasury", {})
    top_providers = payload.get("top_providers", [])
    recent_batches = payload.get("recent_batches", [])
    duplicate_groups = payload.get("duplicate_groups", [])

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Resumen Ejecutivo"
    summary_ws.sheet_view.showGridLines = False
    _merge_title(
        summary_ws,
        "Dashboard Confirming",
        f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')} con resumen operativo, tesoreria y calidad de datos.",
    )

    _metric_card(summary_ws, 2, 6, "Remesas procesadas", stats.get("processed_batches", 0), "DBEAFE")
    _metric_card(summary_ws, 4, 6, "Importe total", stats.get("total_amount", 0.0), "DCFCE7")
    _metric_card(summary_ws, 6, 6, "Incidencias", stats.get("issues_count", 0), "FFEDD5")
    _metric_card(summary_ws, 8, 6, "Duplicados", stats.get("duplicate_invoices_count", 0), "FEF3C7")
    summary_ws["E7"].number_format = CURRENCY_FORMAT

    summary_ws.merge_cells("B10:E10")
    _set_cell(summary_ws["B10"], "Escenario base de tesoreria", font=Font(name="Calibri", size=12, bold=True), fill=SECTION_FILL)
    summary_ws.merge_cells("F10:I10")
    _set_cell(summary_ws["F10"], "Alertas clave", font=Font(name="Calibri", size=12, bold=True), fill=SECTION_FILL)

    treasury_summary = treasury.get("summary", {})
    treasury_rows = [
        ["Saldo inicial", treasury.get("opening_balance", 0.0)],
        ["Reserva minima", treasury.get("reserve_balance", 0.0)],
        ["Pagos planificados", treasury_summary.get("scheduled_total", 0.0)],
        ["Saldo final", treasury_summary.get("final_balance", 0.0)],
    ]
    treasury_end = _write_table(summary_ws, 11, 2, ["Concepto", "Valor"], treasury_rows)
    _format_currency_columns(summary_ws, [3], 12, treasury_end)

    alerts = treasury.get("alerts", [])
    alert_rows = [[alert.get("type", "info").upper(), alert.get("message", "")] for alert in alerts] or [["INFO", "Sin alertas activas"]]
    _write_table(summary_ws, 11, 6, ["Nivel", "Mensaje"], alert_rows)

    recent_rows = [
        [item.get("id"), item.get("name"), item.get("status"), item.get("invoice_count"), item.get("total_amount", 0.0)]
        for item in recent_batches
    ]
    recent_end = _write_table(summary_ws, 18, 2, ["ID", "Lote", "Estado", "Facturas", "Importe"], recent_rows, title="Actividad reciente")
    _format_currency_columns(summary_ws, [6], 20, recent_end)

    provider_rows = [
        [item.get("name"), item.get("cif"), item.get("invoice_count"), item.get("total_amount", 0.0)]
        for item in top_providers
    ]
    providers_end = _write_table(summary_ws, 18, 8, ["Proveedor", "CIF", "Facturas", "Importe"], provider_rows, title="Top proveedores")
    _format_currency_columns(summary_ws, [11], 20, providers_end)
    adjust_column_width(summary_ws)

    volume_ws = workbook.create_sheet("Volumen Mensual")
    volume_ws.sheet_view.showGridLines = False
    monthly_rows = [
        [item.get("full_date"), item.get("name"), item.get("amount", 0.0)] for item in stats.get("monthly_volume", [])
    ]
    volume_end = _write_table(volume_ws, 2, 2, ["Periodo", "Mes", "Importe"], monthly_rows, title="Volumen mensual de confirming")
    _format_currency_columns(volume_ws, [4], 4, volume_end)
    _create_bar_chart(volume_ws, 4, 3, volume_end, "Volumen mensual", "F3", category_col=3)
    adjust_column_width(volume_ws)

    cash_ws = workbook.create_sheet("Tesoreria")
    cash_ws.sheet_view.showGridLines = False
    week_rows = [
        [
            week.get("label"),
            week.get("range"),
            week.get("scheduled_amount", 0.0),
            week.get("delayed_amount", 0.0),
            week.get("scheduled_balance", 0.0),
            week.get("delayed_balance", 0.0),
            week.get("stressed_balance", 0.0),
            week.get("available_after_reserve", 0.0),
        ]
        for week in treasury.get("weeks", [])
    ]
    cash_end = _write_table(
        cash_ws,
        2,
        2,
        ["Semana", "Rango", "Pagos base", "Pagos retrasados", "Saldo base", "Saldo retrasado", "Saldo estresado", "Libre tras reserva"],
        week_rows,
        title="Escenarios de tesoreria",
    )
    _format_currency_columns(cash_ws, [4, 5, 6, 7, 8, 9], 4, cash_end)
    _create_line_chart(cash_ws, [6, 7, 8], 3, cash_end, "Evolucion de saldo", "K3", category_col=2)
    _create_bar_chart(cash_ws, 4, 3, cash_end, "Pagos planificados", "K20", category_col=2)
    adjust_column_width(cash_ws)

    providers_ws = workbook.create_sheet("Proveedores")
    providers_ws.sheet_view.showGridLines = False
    provider_detail_rows = [
        [
            item.get("name"),
            item.get("cif"),
            item.get("invoice_count"),
            item.get("total_amount", 0.0),
            item.get("first_due_date"),
            item.get("last_due_date"),
        ]
        for item in top_providers
    ]
    provider_end = _write_table(
        providers_ws,
        2,
        2,
        ["Proveedor", "CIF", "Facturas", "Importe acumulado", "Primer vencimiento", "Ultimo vencimiento"],
        provider_detail_rows,
        title="Ranking de proveedores por volumen",
    )
    _format_currency_columns(providers_ws, [5], 4, provider_end)
    _create_bar_chart(providers_ws, 5, 3, provider_end, "Top proveedores por importe", "I3", category_col=2)

    exposure_rows = [
        [item.get("name"), item.get("cif"), item.get("invoices"), item.get("amount", 0.0)]
        for item in treasury.get("top_exposures", [])
    ]
    exposure_end = _write_table(
        providers_ws,
        provider_end + 3,
        2,
        ["Proveedor", "CIF", "Facturas proximas", "Exposicion 30 dias"],
        exposure_rows,
        title="Mayor presion de caja (30 dias)",
    )
    _format_currency_columns(providers_ws, [5], provider_end + 5, exposure_end)
    adjust_column_width(providers_ws)

    quality_ws = workbook.create_sheet("Calidad Datos")
    quality_ws.sheet_view.showGridLines = False
    status_rows = [
        [item.get("name"), item.get("value", 0)] for item in stats.get("status_distribution", [])
    ]
    status_end = _write_table(quality_ws, 2, 2, ["Estado", "Facturas"], status_rows, title="Distribucion de calidad")
    _create_pie_chart(quality_ws, 2, 3, 3, status_end, "Calidad de datos", "F3")

    duplicate_rows = [
        [
            group.get("reference"),
            group.get("occurrences", 0),
            group.get("amount", 0.0),
            group.get("total_amount", 0.0),
            ", ".join(f"#{batch_id}" for batch_id in group.get("batch_ids", [])) or "-",
        ]
        for group in duplicate_groups[:15]
    ]
    duplicate_end = _write_table(
        quality_ws,
        status_end + 4,
        2,
        ["Referencia", "Ocurrencias", "Importe unitario", "Importe total", "Lotes"],
        duplicate_rows,
        title="Resumen de duplicados",
    )
    _format_currency_columns(quality_ws, [4, 5], status_end + 6, duplicate_end)
    if duplicate_end > status_end + 5:
        _create_bar_chart(quality_ws, 3, status_end + 5, duplicate_end, "Duplicados por referencia", "H18", category_col=2)
    adjust_column_width(quality_ws)

    activity_ws = workbook.create_sheet("Actividad")
    activity_ws.sheet_view.showGridLines = False
    activity_rows = [
        [
            item.get("id"),
            item.get("name"),
            item.get("status"),
            item.get("created_at"),
            item.get("payment_date"),
            item.get("invoice_count"),
            item.get("total_amount", 0.0),
        ]
        for item in recent_batches
    ]
    activity_end = _write_table(
        activity_ws,
        2,
        2,
        ["ID", "Lote", "Estado", "Creado", "Pago", "Facturas", "Importe"],
        activity_rows,
        title="Seguimiento de lotes recientes",
    )
    _format_currency_columns(activity_ws, [8], 4, activity_end)
    adjust_column_width(activity_ws)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()
